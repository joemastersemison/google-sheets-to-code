#!/usr/bin/env python3
"""
Generate XLSX files for the Google Sheets to Code examples.
Creates fully functional Excel files with data and formulas.
"""

import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_financial_model():
    """Create financial-model.xlsx with loan calculations and investment analysis."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    loan_params = wb.create_sheet("LoanParameters")
    cash_flows = wb.create_sheet("CashFlows")
    loan_calcs = wb.create_sheet("LoanCalculations")
    investment = wb.create_sheet("InvestmentAnalysis")
    
    # Style definitions
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    label_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # LoanParameters Sheet
    loan_params['A1'] = 'Loan Parameters'
    loan_params['A1'].font = Font(bold=True, size=14)
    loan_params.merge_cells('A1:B1')
    
    params_data = [
        ['Annual Interest Rate', 0.05, '5% annual rate'],
        ['Loan Term (Years)', 30, '30-year mortgage'],
        ['Loan Principal', 500000, '$500,000 loan'],
        ['', '', ''],
        ['Investment Parameters', '', ''],
        ['Investment Annual Rate', 0.08, '8% annual return'],
        ['Investment Period (Years)', 20, '20-year investment'],
        ['Monthly Contribution', 1000, '$1,000/month'],
        ['Initial Investment', 10000, '$10,000 initial'],
        ['', '', ''],
        ['NPV/IRR Parameters', '', ''],
        ['Discount Rate', 0.10, '10% discount rate'],
        ['Number of Periods', 10, '10 periods'],
        ['Payment Amount', 5000, '$5,000 payment'],
        ['Present Value', 40000, '$40,000 PV'],
        ['Future Value', 0, '$0 FV']
    ]
    
    for i, row in enumerate(params_data, start=2):
        if row[0] in ['Investment Parameters', 'NPV/IRR Parameters']:
            # For section headers, set value in first cell then merge
            cell = loan_params.cell(row=i, column=1, value=row[0])
            cell.font = Font(bold=True, size=12)
            loan_params.merge_cells(f'A{i}:C{i}')
        else:
            # For regular rows, set all values
            for j, value in enumerate(row, start=1):
                cell = loan_params.cell(row=i, column=j, value=value)
                if j == 1 and value and value not in ['']:
                    cell.font = label_font
    
    # CashFlows Sheet
    cash_flows['A1'] = 'Period'
    cash_flows['B1'] = 'Cash Flow (NPV)'
    cash_flows['C1'] = 'Cash Flow (IRR)'
    
    for cell in ['A1', 'B1', 'C1']:
        cash_flows[cell].font = header_font
        cash_flows[cell].fill = header_fill
    
    cash_flow_data = [
        [0, 0, -100000],
        [1, 15000, 15000],
        [2, 20000, 20000],
        [3, 25000, 25000],
        [4, 30000, 30000],
        [5, 35000, 35000],
        [6, 40000, 40000],
        [7, 45000, 45000],
        [8, 50000, 50000],
        [9, 55000, 55000],
        [10, 60000, 60000]
    ]
    
    for i, row in enumerate(cash_flow_data, start=2):
        for j, value in enumerate(row, start=1):
            cash_flows.cell(row=i, column=j, value=value)
    
    # LoanCalculations Sheet
    loan_calcs['A1'] = 'Loan Analysis'
    loan_calcs['A1'].font = Font(bold=True, size=14)
    loan_calcs.merge_cells('A1:B1')
    
    loan_calc_formulas = [
        ['Monthly Payment', '=PMT(LoanParameters!B2/12, LoanParameters!B3*12, -LoanParameters!B4)'],
        ['Total Amount Paid', '=B2*LoanParameters!B3*12'],
        ['Total Interest Paid', '=B3-LoanParameters!B4'],
        ['Effective Annual Rate', '=RATE(LoanParameters!B14, -LoanParameters!B15, LoanParameters!B16, -LoanParameters!B17)*12']
    ]
    
    for i, (label, formula) in enumerate(loan_calc_formulas, start=2):
        loan_calcs.cell(row=i, column=1, value=label).font = label_font
        loan_calcs.cell(row=i, column=2, value=formula)
    
    # Amortization Schedule Header
    loan_calcs['A7'] = 'Amortization Schedule'
    loan_calcs['A7'].font = Font(bold=True, size=12)
    loan_calcs.merge_cells('A7:D7')
    
    loan_calcs['A8'] = 'Month'
    loan_calcs['B8'] = 'Principal'
    loan_calcs['C8'] = 'Interest'
    loan_calcs['D8'] = 'Balance'
    
    for col in ['A8', 'B8', 'C8', 'D8']:
        loan_calcs[col].font = header_font
        loan_calcs[col].fill = header_fill
    
    # Add first few rows of amortization
    loan_calcs['A9'] = 1
    loan_calcs['B9'] = '=B2-C9'
    loan_calcs['C9'] = '=LoanParameters!B4*LoanParameters!B2/12'
    loan_calcs['D9'] = '=LoanParameters!B4-B9'
    
    loan_calcs['A10'] = 2
    loan_calcs['B10'] = '=B2-C10'
    loan_calcs['C10'] = '=D9*LoanParameters!B2/12'
    loan_calcs['D10'] = '=D9-B10'
    
    # InvestmentAnalysis Sheet
    investment['A1'] = 'Investment Growth'
    investment['A1'].font = Font(bold=True, size=14)
    investment.merge_cells('A1:B1')
    
    investment_formulas = [
        ['Future Value (Lump Sum)', '=FV(LoanParameters!B7, LoanParameters!B8, 0, -LoanParameters!B10)'],
        ['Future Value (With Contributions)', '=FV(LoanParameters!B7/12, LoanParameters!B8*12, -LoanParameters!B9, -LoanParameters!B10)'],
        ['Present Value Needed', '=PV(LoanParameters!B7, LoanParameters!B8, 0, -1000000)'],
        ['Monthly Contribution Needed', '=PMT(LoanParameters!B7/12, LoanParameters!B8*12, -LoanParameters!B10, 1000000)'],
        ['', ''],
        ['Project Analysis', ''],
        ['Net Present Value', '=NPV(LoanParameters!B13, CashFlows!B3:B12)'],
        ['Internal Rate of Return', '=IRR(CashFlows!C2:C12)'],
        ['Profitability Index', '=1 + B8/ABS(CashFlows!C2)']
    ]
    
    for i, row in enumerate(investment_formulas, start=2):
        if len(row) == 2:
            label, formula = row
            investment.cell(row=i, column=1, value=label).font = label_font if label else None
            investment.cell(row=i, column=2, value=formula)
        elif len(row) == 1:
            investment.cell(row=i, column=1, value=row[0]).font = Font(bold=True, size=12)
    
    # Sensitivity Analysis
    investment['A13'] = 'Sensitivity Analysis'
    investment['A13'].font = Font(bold=True, size=12)
    investment.merge_cells('A13:C13')
    
    investment['A14'] = 'Discount Rate'
    investment['B14'] = 'NPV'
    investment['C14'] = 'Decision'
    
    for cell in ['A14', 'B14', 'C14']:
        investment[cell].font = header_font
        investment[cell].fill = header_fill
    
    sensitivity_data = [
        [0.05, '=NPV(0.05, CashFlows!B3:B12)', '=IF(B15>0, "Accept", "Reject")'],
        [0.10, '=NPV(0.10, CashFlows!B3:B12)', '=IF(B16>0, "Accept", "Reject")'],
        [0.15, '=NPV(0.15, CashFlows!B3:B12)', '=IF(B17>0, "Accept", "Reject")'],
        [0.20, '=NPV(0.20, CashFlows!B3:B12)', '=IF(B18>0, "Accept", "Reject")']
    ]
    
    for i, row in enumerate(sensitivity_data, start=15):
        for j, value in enumerate(row, start=1):
            investment.cell(row=i, column=j, value=value)
    
    # Adjust column widths
    for sheet in [loan_params, cash_flows, loan_calcs, investment]:
        for col in range(1, 5):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Save the workbook
    wb.save('financial-model.xlsx')
    print("✅ Created financial-model.xlsx")

def create_data_analysis():
    """Create data-analysis.xlsx with statistical analysis examples."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    raw_data = wb.create_sheet("RawData")
    parameters = wb.create_sheet("Parameters")
    analysis = wb.create_sheet("Analysis")
    summary = wb.create_sheet("Summary")
    
    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    label_font = Font(bold=True)
    
    # RawData Sheet - Generate sample sales data
    headers = ['Sales Amount', 'Region', 'Product', 'Date', 'Customer ID']
    for i, header in enumerate(headers, start=1):
        cell = raw_data.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    regions = ['North', 'South', 'East', 'West']
    products = ['Widget A', 'Widget B', 'Widget C']
    
    # Generate 100 rows of sample data
    random.seed(42)  # For reproducibility
    base_date = datetime(2024, 1, 1)
    
    for i in range(2, 102):
        raw_data.cell(row=i, column=1, value=random.randint(800, 4500))  # Sales Amount
        raw_data.cell(row=i, column=2, value=random.choice(regions))  # Region
        raw_data.cell(row=i, column=3, value=random.choice(products))  # Product
        raw_data.cell(row=i, column=4, value=(base_date + timedelta(days=i-2)).strftime('%m/%d/%Y'))  # Date
        raw_data.cell(row=i, column=5, value=f'C{i-1:03d}')  # Customer ID
    
    # Parameters Sheet
    parameters['A1'] = 'Analysis Parameters'
    parameters['A1'].font = Font(bold=True, size=14)
    parameters.merge_cells('A1:B1')
    
    params_data = [
        ['Confidence Level', 0.95, '95% confidence'],
        ['Significance Threshold', 0.05, '5% significance'],
        ['Moving Average Window', 7, '7-day window'],
        ['Outlier Threshold (σ)', 3, '3 standard deviations'],
        ['Minimum Sample Size', 30, 'Min 30 samples'],
        ['Top N Items', 10, 'Top 10 analysis'],
        ['Sales Target', 2500, '$2,500 target']
    ]
    
    for i, row in enumerate(params_data, start=2):
        for j, value in enumerate(row, start=1):
            cell = parameters.cell(row=i, column=j, value=value)
            if j == 1:
                cell.font = label_font
    
    # Analysis Sheet
    analysis['A1'] = 'Basic Statistics'
    analysis['A1'].font = Font(bold=True, size=14)
    analysis.merge_cells('A1:B1')
    
    basic_stats = [
        ['Mean Sales', '=AVERAGE(RawData!A2:A101)'],
        ['Median Sales', '=MEDIAN(RawData!A2:A101)'],
        ['Standard Deviation', '=STDEV(RawData!A2:A101)'],
        ['Variance', '=VAR(RawData!A2:A101)'],
        ['Min Sales', '=MIN(RawData!A2:A101)'],
        ['Max Sales', '=MAX(RawData!A2:A101)'],
        ['Range', '=B7-B6'],
        ['Count', '=COUNT(RawData!A2:A101)'],
        ['Count Above Target', '=COUNTIF(RawData!A2:A101, ">"&Parameters!B8)']
    ]
    
    for i, (label, formula) in enumerate(basic_stats, start=2):
        analysis.cell(row=i, column=1, value=label).font = label_font
        analysis.cell(row=i, column=2, value=formula)
    
    # Percentile Analysis
    analysis['A12'] = 'Percentile Analysis'
    analysis['A12'].font = Font(bold=True, size=12)
    analysis.merge_cells('A12:B12')
    
    percentiles = [
        ['25th Percentile', '=PERCENTILE(RawData!A2:A101, 0.25)'],
        ['50th Percentile', '=PERCENTILE(RawData!A2:A101, 0.50)'],
        ['75th Percentile', '=PERCENTILE(RawData!A2:A101, 0.75)'],
        ['90th Percentile', '=PERCENTILE(RawData!A2:A101, 0.90)'],
        ['Interquartile Range', '=B15-B13']
    ]
    
    for i, (label, formula) in enumerate(percentiles, start=13):
        analysis.cell(row=i, column=1, value=label).font = label_font
        analysis.cell(row=i, column=2, value=formula)
    
    # Regional Analysis
    analysis['A24'] = 'Regional Analysis'
    analysis['A24'].font = Font(bold=True, size=12)
    analysis.merge_cells('A24:D24')
    
    analysis['A25'] = 'Region'
    analysis['B25'] = 'Count'
    analysis['C25'] = 'Average'
    analysis['D25'] = 'Total'
    
    for cell in ['A25', 'B25', 'C25', 'D25']:
        analysis[cell].font = header_font
        analysis[cell].fill = header_fill
    
    for i, region in enumerate(regions, start=26):
        analysis.cell(row=i, column=1, value=region)
        analysis.cell(row=i, column=2, value=f'=COUNTIF(RawData!B2:B101, "{region}")')
        analysis.cell(row=i, column=3, value=f'=AVERAGEIF(RawData!B2:B101, "{region}", RawData!A2:A101)')
        analysis.cell(row=i, column=4, value=f'=SUMIF(RawData!B2:B101, "{region}", RawData!A2:A101)')
    
    # Summary Sheet
    summary['A1'] = 'Executive Summary'
    summary['A1'].font = Font(bold=True, size=14)
    summary.merge_cells('A1:B1')
    
    summary_data = [
        ['Total Sales', '=SUM(RawData!A2:A101)'],
        ['Average Daily Sales', '=Analysis!B2'],
        ['Best Performing Region', '=INDEX(Analysis!A26:A29, MATCH(MAX(Analysis!D26:D29), Analysis!D26:D29, 0))'],
        ['Sales Above Target (%)', '=Analysis!B10/Analysis!B9*100'],
        ['', ''],
        ['Data Quality Report', ''],
        ['Total Records', '=Analysis!B9'],
        ['Data Completeness (%)', '=COUNTA(RawData!A2:A101)/100*100'],
        ['', ''],
        ['Statistical Tests', ''],
        ['Sample Size Adequate', '=IF(Analysis!B9>=Parameters!B6, "Yes", "No")']
    ]
    
    for i, row in enumerate(summary_data, start=2):
        if len(row) == 2:
            label, formula = row
            cell = summary.cell(row=i, column=1, value=label)
            if label and label not in ['Data Quality Report', 'Statistical Tests']:
                cell.font = label_font
            elif label in ['Data Quality Report', 'Statistical Tests']:
                cell.font = Font(bold=True, size=12)
            summary.cell(row=i, column=2, value=formula)
        elif len(row) == 1:
            summary.cell(row=i, column=1, value=row[0]).font = Font(bold=True, size=12)
    
    # Adjust column widths
    for sheet in [raw_data, parameters, analysis, summary]:
        for col in range(1, 6):
            sheet.column_dimensions[get_column_letter(col)].width = 18
    
    # Save the workbook
    wb.save('data-analysis.xlsx')
    print("✅ Created data-analysis.xlsx")

def create_inventory_tracking():
    """Create inventory-tracking.xlsx with inventory management system."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    products = wb.create_sheet("Products")
    transactions = wb.create_sheet("Transactions")
    settings = wb.create_sheet("Settings")
    current_inv = wb.create_sheet("CurrentInventory")
    alerts = wb.create_sheet("Alerts")
    reports = wb.create_sheet("Reports")
    
    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    label_font = Font(bold=True)
    
    # Products Sheet
    product_headers = ['SKU', 'Product Name', 'Unit Cost', 'Lead Time (days)', 
                      'Safety Stock', 'Annual Demand', 'Supplier', 'Category']
    
    for i, header in enumerate(product_headers, start=1):
        cell = products.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    product_data = [
        ['SKU001', 'Widget Alpha', 25.50, 7, 50, 2400, 'Supplier A', 'Electronics'],
        ['SKU002', 'Widget Beta', 45.00, 14, 30, 1800, 'Supplier B', 'Electronics'],
        ['SKU003', 'Widget Gamma', 12.75, 3, 100, 5000, 'Supplier A', 'Accessories'],
        ['SKU004', 'Widget Delta', 89.99, 21, 20, 600, 'Supplier C', 'Premium'],
        ['SKU005', 'Widget Epsilon', 5.25, 5, 200, 10000, 'Supplier B', 'Consumables'],
        ['SKU006', 'Widget Zeta', 150.00, 30, 10, 200, 'Supplier D', 'Premium'],
        ['SKU007', 'Widget Eta', 35.00, 10, 40, 1500, 'Supplier A', 'Electronics'],
        ['SKU008', 'Widget Theta', 18.50, 7, 75, 3000, 'Supplier C', 'Accessories'],
        ['SKU009', 'Widget Iota', 62.00, 14, 25, 900, 'Supplier B', 'Electronics'],
        ['SKU010', 'Widget Kappa', 8.99, 3, 150, 7500, 'Supplier A', 'Consumables']
    ]
    
    for i, row in enumerate(product_data, start=2):
        for j, value in enumerate(row, start=1):
            products.cell(row=i, column=j, value=value)
    
    # Transactions Sheet
    trans_headers = ['Date', 'Product SKU', 'Transaction Type', 'Quantity', 
                    'Unit Price', 'Reference Number', 'Notes']
    
    for i, header in enumerate(trans_headers, start=1):
        cell = transactions.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Generate sample transactions
    base_date = datetime(2024, 1, 1)
    trans_data = []
    
    # Initial stock for all products
    for sku in ['SKU001', 'SKU002', 'SKU003', 'SKU004', 'SKU005']:
        trans_data.append([
            base_date.strftime('%m/%d/%Y'),
            sku,
            'IN',
            random.randint(100, 300),
            0,  # Will be looked up
            'PO-2024-001',
            'Initial stock'
        ])
    
    # Generate random transactions
    for i in range(50):
        date = (base_date + timedelta(days=i//2)).strftime('%m/%d/%Y')
        sku = random.choice(['SKU001', 'SKU002', 'SKU003', 'SKU004', 'SKU005'])
        trans_type = random.choice(['IN', 'OUT', 'OUT', 'OUT'])  # More outgoing
        quantity = random.randint(5, 50) if trans_type == 'OUT' else random.randint(50, 200)
        ref = f'SO-2024-{i+1:03d}' if trans_type == 'OUT' else f'PO-2024-{i//10+2:03d}'
        
        trans_data.append([
            date,
            sku,
            trans_type,
            quantity,
            0,  # Will be looked up or calculated
            ref,
            'Customer order' if trans_type == 'OUT' else 'Restock'
        ])
    
    for i, row in enumerate(trans_data, start=2):
        for j, value in enumerate(row, start=1):
            transactions.cell(row=i, column=j, value=value)
    
    # Settings Sheet
    settings['A1'] = 'Inventory Settings'
    settings['A1'].font = Font(bold=True, size=14)
    settings.merge_cells('A1:B1')
    
    settings_data = [
        ['Lead Time Multiplier', 1.5, 'Safety factor for lead time'],
        ['Ordering Cost per Order', 50, 'Fixed cost per order'],
        ['Holding Cost Rate (annual %)', 0.20, '20% annual holding cost'],
        ['Service Level Target', 0.95, '95% service level'],
        ['Review Period (days)', 7, 'Weekly review'],
        ['Min Order Quantity', 10, 'Minimum order size'],
        ['Max Order Quantity', 1000, 'Maximum order size'],
        ['Critical Stock Level (%)', 0.25, '25% of reorder point'],
        ['Overstock Threshold (days)', 90, '90 days of supply']
    ]
    
    for i, row in enumerate(settings_data, start=2):
        for j, value in enumerate(row, start=1):
            cell = settings.cell(row=i, column=j, value=value)
            if j == 1:
                cell.font = label_font
    
    # CurrentInventory Sheet
    inv_headers = ['SKU', 'Current Stock', 'Reorder Point', 'Status', 'Stock Value', 
                  'Days of Supply', 'Turnover', 'ABC Class', 'Risk', 'EOQ']
    
    for i, header in enumerate(inv_headers, start=1):
        cell = current_inv.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add formulas for first few SKUs
    for i, sku in enumerate(['SKU001', 'SKU002', 'SKU003', 'SKU004', 'SKU005'], start=2):
        current_inv.cell(row=i, column=1, value=sku)
        current_inv.cell(row=i, column=2, value=f'=SUMIFS(Transactions!D:D, Transactions!B:B, A{i}, Transactions!C:C, "IN") - SUMIFS(Transactions!D:D, Transactions!B:B, A{i}, Transactions!C:C, "OUT")')
        current_inv.cell(row=i, column=3, value=f'=VLOOKUP(A{i}, Products!A:E, 4, FALSE) * Settings!B2 * (VLOOKUP(A{i}, Products!A:F, 6, FALSE)/365) + VLOOKUP(A{i}, Products!A:E, 5, FALSE)')
        current_inv.cell(row=i, column=4, value=f'=IF(B{i}<=C{i}*Settings!B9, "CRITICAL", IF(B{i}<=C{i}, "REORDER", "OK"))')
        current_inv.cell(row=i, column=5, value=f'=B{i} * VLOOKUP(A{i}, Products!A:C, 3, FALSE)')
        current_inv.cell(row=i, column=6, value=f'=IF(SUMIFS(Transactions!D:D, Transactions!B:B, A{i}, Transactions!C:C, "OUT")/30>0, B{i}/(SUMIFS(Transactions!D:D, Transactions!B:B, A{i}, Transactions!C:C, "OUT")/30), 999)')
        current_inv.cell(row=i, column=7, value=f'=IF(B{i}>0, SUMIFS(Transactions!D:D, Transactions!B:B, A{i}, Transactions!C:C, "OUT")/B{i}*365/30, 0)')
        current_inv.cell(row=i, column=8, value=f'=IF(E{i}/SUM(E:E)>0.7, "A", IF(E{i}/SUM(E:E)>0.2, "B", "C"))')
        current_inv.cell(row=i, column=9, value=f'=IF(B{i}<C{i}*0.5, "HIGH", IF(B{i}<C{i}, "MEDIUM", "LOW"))')
        current_inv.cell(row=i, column=10, value=f'=SQRT(2*VLOOKUP(A{i}, Products!A:F, 6, FALSE)*Settings!B3/(VLOOKUP(A{i}, Products!A:C, 3, FALSE)*Settings!B4))')
    
    # Alerts Sheet
    alerts['A1'] = 'Alert Type'
    alerts['B1'] = 'Product SKU'
    alerts['C1'] = 'Message'
    alerts['D1'] = 'Priority'
    alerts['E1'] = 'Action Required'
    alerts['F1'] = 'Date'
    
    for col in range(1, 7):
        cell = alerts.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    alert_types = ['Stock Level', 'Reorder Point', 'Overstock', 'Slow Moving']
    
    for i, alert_type in enumerate(alert_types, start=2):
        alerts.cell(row=i, column=1, value=alert_type)
        
        if alert_type == 'Stock Level':
            alerts.cell(row=i, column=2, value='=IF(COUNTIF(CurrentInventory!D:D, "CRITICAL")>0, INDEX(CurrentInventory!A:A, MATCH("CRITICAL", CurrentInventory!D:D, 0)), "")')
            alerts.cell(row=i, column=3, value='=IF(B2<>"", "Critical stock level - immediate reorder required", "")')
            alerts.cell(row=i, column=4, value='=IF(B2<>"", "HIGH", "")')
        elif alert_type == 'Reorder Point':
            alerts.cell(row=i, column=2, value='=IF(COUNTIF(CurrentInventory!D:D, "REORDER")>0, INDEX(CurrentInventory!A:A, MATCH("REORDER", CurrentInventory!D:D, 0)), "")')
            alerts.cell(row=i, column=3, value='=IF(B3<>"", "Stock at reorder point - place order", "")')
            alerts.cell(row=i, column=4, value='=IF(B3<>"", "MEDIUM", "")')
    
    # Reports Sheet
    reports['A1'] = 'Inventory Dashboard'
    reports['A1'].font = Font(bold=True, size=14)
    reports.merge_cells('A1:B1')
    
    report_data = [
        ['Total Inventory Value', '=SUM(CurrentInventory!E:E)'],
        ['Number of SKUs', '=COUNTA(CurrentInventory!A:A)-1'],
        ['Items Below Reorder Point', '=COUNTIF(CurrentInventory!D:D, "REORDER") + COUNTIF(CurrentInventory!D:D, "CRITICAL")'],
        ['Items Out of Stock', '=COUNTIF(CurrentInventory!B:B, 0)'],
        ['Critical Stock Items', '=COUNTIF(CurrentInventory!D:D, "CRITICAL")'],
        ['', ''],
        ['Performance Metrics', ''],
        ['Average Turnover Rate', '=AVERAGE(CurrentInventory!G:G)'],
        ['Fill Rate (%)', '=(1-B5/B3)*100'],
        ['Service Level (%)', '=COUNTIF(CurrentInventory!D:D, "OK")/B3*100']
    ]
    
    for i, row in enumerate(report_data, start=2):
        if len(row) == 2:
            label, formula = row
            cell = reports.cell(row=i, column=1, value=label)
            if label and label not in ['Performance Metrics']:
                cell.font = label_font
            elif label == 'Performance Metrics':
                cell.font = Font(bold=True, size=12)
            reports.cell(row=i, column=2, value=formula)
        elif len(row) == 1:
            reports.cell(row=i, column=1, value=row[0]).font = Font(bold=True, size=12)
    
    # Adjust column widths
    for sheet in [products, transactions, settings, current_inv, alerts, reports]:
        for col in range(1, 11):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Save the workbook
    wb.save('inventory-tracking.xlsx')
    print("✅ Created inventory-tracking.xlsx")

def main():
    """Generate all example XLSX files."""
    print("🚀 Generating XLSX files for examples...")
    print("")
    
    # Check if openpyxl is installed
    try:
        import openpyxl
    except ImportError:
        print("❌ Error: openpyxl is not installed.")
        print("Please install it with: pip install openpyxl")
        return
    
    # Change to examples directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    # Generate each file
    try:
        create_financial_model()
        create_data_analysis()
        create_inventory_tracking()
        
        print("")
        print("✨ All XLSX files created successfully!")
        print("")
        print("You can now:")
        print("1. Upload these files to Google Sheets (File → Import)")
        print("2. Share the sheets publicly")
        print("3. Update the JSON configuration files with the URLs")
        print("4. Run the converter to generate code")
        
    except Exception as e:
        print(f"❌ Error creating files: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()